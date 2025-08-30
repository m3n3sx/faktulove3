"""
Data export and import service for FaktuLove system.
Provides comprehensive data export/import capabilities with Polish formatting and validation.
"""

import csv
import json
import io
import zipfile
import tempfile
from datetime import datetime, date
from decimal import Decimal
from typing import Dict, List, Any, Optional, Union, BinaryIO
import logging
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    Workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    colors = None

from django.http import HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from django.contrib.auth.models import User

from faktury.models import Faktura, Kontrahent, Firma, PozycjaFaktury, Produkt

logger = logging.getLogger(__name__)


class DataExportImportService:
    """
    Comprehensive data export and import service with Polish business formatting.
    Supports multiple formats: PDF, Excel, CSV, JSON with progress tracking.
    """
    
    # Polish column headers for exports
    INVOICE_HEADERS_PL = {
        'numer': 'Numer faktury',
        'typ_dokumentu': 'Typ dokumentu',
        'data_wystawienia': 'Data wystawienia',
        'data_sprzedazy': 'Data sprzedaży',
        'termin_platnosci': 'Termin płatności',
        'nabywca__nazwa': 'Nabywca',
        'nabywca__nip': 'NIP nabywcy',
        'nabywca__miejscowosc': 'Miasto nabywcy',
        'sprzedawca__nazwa': 'Sprzedawca',
        'sprzedawca__nip': 'NIP sprzedawcy',
        'status': 'Status',
        'sposob_platnosci': 'Sposób płatności',
        'waluta': 'Waluta',
        'kwota_netto': 'Kwota netto',
        'kwota_vat': 'Kwota VAT',
        'kwota_brutto': 'Kwota brutto',
        'uwagi': 'Uwagi'
    }
    
    COMPANY_HEADERS_PL = {
        'nazwa': 'Nazwa',
        'nip': 'NIP',
        'regon': 'REGON',
        'ulica': 'Ulica',
        'numer_domu': 'Numer domu',
        'numer_mieszkania': 'Numer mieszkania',
        'kod_pocztowy': 'Kod pocztowy',
        'miejscowosc': 'Miejscowość',
        'kraj': 'Kraj',
        'czy_firma': 'Typ',
        'email': 'Email',
        'telefon': 'Telefon',
        'dodatkowy_opis': 'Opis'
    }
    
    PRODUCT_HEADERS_PL = {
        'nazwa': 'Nazwa produktu',
        'jednostka': 'Jednostka',
        'cena_netto': 'Cena netto',
        'vat': 'Stawka VAT',
        'cena_brutto': 'Cena brutto'
    }
    
    def __init__(self):
        self.progress_callbacks = {}
    
    def export_invoices(
        self,
        user: User,
        format_type: str = 'excel',
        filters: Dict[str, Any] = None,
        progress_callback_id: str = None
    ) -> Union[HttpResponse, Dict[str, Any]]:
        """
        Export invoices in specified format with Polish formatting.
        
        Args:
            user: User object for filtering data
            format_type: 'excel', 'csv', 'pdf', 'json'
            filters: Optional filters to apply
            progress_callback_id: ID for progress tracking
            
        Returns:
            HttpResponse with file or dict with export info
        """
        try:
            self._update_progress(progress_callback_id, 10, "Pobieranie danych faktur...")
            
            # Get invoices queryset
            queryset = Faktura.objects.filter(user=user).select_related(
                'nabywca', 'sprzedawca'
            ).prefetch_related('pozycjafaktury_set')
            
            # Apply filters if provided
            if filters:
                queryset = self._apply_invoice_filters(queryset, filters)
            
            self._update_progress(progress_callback_id, 30, "Przetwarzanie danych...")
            
            # Prepare data
            data = self._prepare_invoice_data(queryset)
            
            self._update_progress(progress_callback_id, 60, f"Generowanie pliku {format_type.upper()}...")
            
            # Generate export based on format
            if format_type == 'excel':
                response = self._export_invoices_excel(data, user)
            elif format_type == 'csv':
                response = self._export_invoices_csv(data, user)
            elif format_type == 'pdf':
                response = self._export_invoices_pdf(data, user)
            elif format_type == 'json':
                response = self._export_invoices_json(data, user)
            else:
                raise ValueError(f"Nieobsługiwany format: {format_type}")
            
            self._update_progress(progress_callback_id, 100, "Eksport zakończony pomyślnie")
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting invoices: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd eksportu: {str(e)}")
            raise
    
    def export_companies(
        self,
        user: User,
        format_type: str = 'excel',
        filters: Dict[str, Any] = None,
        progress_callback_id: str = None
    ) -> Union[HttpResponse, Dict[str, Any]]:
        """Export companies in specified format."""
        try:
            self._update_progress(progress_callback_id, 10, "Pobieranie danych kontrahentów...")
            
            # Get companies queryset
            queryset = Kontrahent.objects.filter(user=user)
            
            # Apply filters if provided
            if filters:
                queryset = self._apply_company_filters(queryset, filters)
            
            self._update_progress(progress_callback_id, 30, "Przetwarzanie danych...")
            
            # Prepare data
            data = self._prepare_company_data(queryset)
            
            self._update_progress(progress_callback_id, 60, f"Generowanie pliku {format_type.upper()}...")
            
            # Generate export based on format
            if format_type == 'excel':
                response = self._export_companies_excel(data, user)
            elif format_type == 'csv':
                response = self._export_companies_csv(data, user)
            elif format_type == 'pdf':
                response = self._export_companies_pdf(data, user)
            elif format_type == 'json':
                response = self._export_companies_json(data, user)
            else:
                raise ValueError(f"Nieobsługiwany format: {format_type}")
            
            self._update_progress(progress_callback_id, 100, "Eksport zakończony pomyślnie")
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting companies: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd eksportu: {str(e)}")
            raise
    
    def export_products(
        self,
        user: User,
        format_type: str = 'excel',
        progress_callback_id: str = None
    ) -> Union[HttpResponse, Dict[str, Any]]:
        """Export products in specified format."""
        try:
            self._update_progress(progress_callback_id, 10, "Pobieranie danych produktów...")
            
            # Get products queryset
            queryset = Produkt.objects.filter(user=user)
            
            self._update_progress(progress_callback_id, 30, "Przetwarzanie danych...")
            
            # Prepare data
            data = self._prepare_product_data(queryset)
            
            self._update_progress(progress_callback_id, 60, f"Generowanie pliku {format_type.upper()}...")
            
            # Generate export based on format
            if format_type == 'excel':
                response = self._export_products_excel(data, user)
            elif format_type == 'csv':
                response = self._export_products_csv(data, user)
            elif format_type == 'json':
                response = self._export_products_json(data, user)
            else:
                raise ValueError(f"Nieobsługiwany format: {format_type}")
            
            self._update_progress(progress_callback_id, 100, "Eksport zakończony pomyślnie")
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting products: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd eksportu: {str(e)}")
            raise
    
    def create_backup(
        self,
        user: User,
        include_files: bool = False,
        progress_callback_id: str = None
    ) -> HttpResponse:
        """Create complete data backup as ZIP file."""
        try:
            self._update_progress(progress_callback_id, 5, "Rozpoczynanie tworzenia kopii zapasowej...")
            
            # Create temporary file for ZIP
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
            
            with zipfile.ZipFile(temp_file.name, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Export invoices
                self._update_progress(progress_callback_id, 20, "Eksportowanie faktur...")
                invoices_data = self._prepare_invoice_data(
                    Faktura.objects.filter(user=user).select_related('nabywca', 'sprzedawca')
                )
                zip_file.writestr('faktury.json', json.dumps(invoices_data, ensure_ascii=False, indent=2))
                
                # Export companies
                self._update_progress(progress_callback_id, 40, "Eksportowanie kontrahentów...")
                companies_data = self._prepare_company_data(Kontrahent.objects.filter(user=user))
                zip_file.writestr('kontrahenci.json', json.dumps(companies_data, ensure_ascii=False, indent=2))
                
                # Export products
                self._update_progress(progress_callback_id, 60, "Eksportowanie produktów...")
                products_data = self._prepare_product_data(Produkt.objects.filter(user=user))
                zip_file.writestr('produkty.json', json.dumps(products_data, ensure_ascii=False, indent=2))
                
                # Add metadata
                self._update_progress(progress_callback_id, 80, "Dodawanie metadanych...")
                metadata = {
                    'created_at': timezone.now().isoformat(),
                    'user_id': user.id,
                    'username': user.username,
                    'version': '1.0',
                    'description': 'Kopia zapasowa danych FaktuLove'
                }
                zip_file.writestr('metadata.json', json.dumps(metadata, ensure_ascii=False, indent=2))
            
            self._update_progress(progress_callback_id, 100, "Kopia zapasowa utworzona pomyślnie")
            
            # Prepare response
            with open(temp_file.name, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename="backup_{user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'
                
            return response
            
        except Exception as e:
            logger.error(f"Error creating backup: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd tworzenia kopii zapasowej: {str(e)}")
            raise
    
    def import_data(
        self,
        user: User,
        file_data: BinaryIO,
        data_type: str,
        format_type: str,
        progress_callback_id: str = None
    ) -> Dict[str, Any]:
        """
        Import data from file with validation and progress tracking.
        
        Args:
            user: User object
            file_data: File data to import
            data_type: 'invoices', 'companies', 'products'
            format_type: 'csv', 'excel', 'json'
            progress_callback_id: ID for progress tracking
            
        Returns:
            Dictionary with import results and statistics
        """
        try:
            self._update_progress(progress_callback_id, 10, "Rozpoczynanie importu danych...")
            
            # Parse file based on format
            if format_type == 'csv':
                data = self._parse_csv_file(file_data)
            elif format_type == 'excel':
                data = self._parse_excel_file(file_data)
            elif format_type == 'json':
                data = self._parse_json_file(file_data)
            else:
                raise ValueError(f"Nieobsługiwany format: {format_type}")
            
            self._update_progress(progress_callback_id, 30, "Walidacja danych...")
            
            # Validate data
            validation_results = self._validate_import_data(data, data_type)
            
            if validation_results['errors']:
                return {
                    'success': False,
                    'errors': validation_results['errors'],
                    'warnings': validation_results['warnings']
                }
            
            self._update_progress(progress_callback_id, 60, "Importowanie danych...")
            
            # Import data
            import_results = self._import_validated_data(user, data, data_type, progress_callback_id)
            
            self._update_progress(progress_callback_id, 100, "Import zakończony pomyślnie")
            
            return {
                'success': True,
                'imported_count': import_results['imported_count'],
                'updated_count': import_results['updated_count'],
                'skipped_count': import_results['skipped_count'],
                'warnings': validation_results['warnings']
            }
            
        except Exception as e:
            logger.error(f"Error importing data: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd importu: {str(e)}")
            return {
                'success': False,
                'errors': [str(e)]
            }
    
    def restore_backup(
        self,
        user: User,
        backup_file: BinaryIO,
        progress_callback_id: str = None
    ) -> Dict[str, Any]:
        """Restore data from backup ZIP file."""
        try:
            self._update_progress(progress_callback_id, 10, "Rozpakowywanie kopii zapasowej...")
            
            with zipfile.ZipFile(backup_file, 'r') as zip_file:
                # Check if it's a valid backup
                if 'metadata.json' not in zip_file.namelist():
                    raise ValueError("Nieprawidłowy plik kopii zapasowej")
                
                # Read metadata
                metadata = json.loads(zip_file.read('metadata.json').decode('utf-8'))
                
                results = {
                    'success': True,
                    'metadata': metadata,
                    'imported': {}
                }
                
                # Restore invoices
                if 'faktury.json' in zip_file.namelist():
                    self._update_progress(progress_callback_id, 30, "Przywracanie faktur...")
                    invoices_data = json.loads(zip_file.read('faktury.json').decode('utf-8'))
                    results['imported']['invoices'] = self._restore_invoices(user, invoices_data)
                
                # Restore companies
                if 'kontrahenci.json' in zip_file.namelist():
                    self._update_progress(progress_callback_id, 60, "Przywracanie kontrahentów...")
                    companies_data = json.loads(zip_file.read('kontrahenci.json').decode('utf-8'))
                    results['imported']['companies'] = self._restore_companies(user, companies_data)
                
                # Restore products
                if 'produkty.json' in zip_file.namelist():
                    self._update_progress(progress_callback_id, 80, "Przywracanie produktów...")
                    products_data = json.loads(zip_file.read('produkty.json').decode('utf-8'))
                    results['imported']['products'] = self._restore_products(user, products_data)
                
                self._update_progress(progress_callback_id, 100, "Przywracanie zakończone pomyślnie")
                
                return results
                
        except Exception as e:
            logger.error(f"Error restoring backup: {str(e)}")
            self._update_progress(progress_callback_id, 0, f"Błąd przywracania: {str(e)}")
            return {
                'success': False,
                'errors': [str(e)]
            }
    
    def get_export_progress(self, progress_id: str) -> Dict[str, Any]:
        """Get export/import progress status."""
        return self.progress_callbacks.get(progress_id, {
            'progress': 0,
            'message': 'Nieznany proces',
            'completed': False
        })
    
    # Private helper methods
    
    def _update_progress(self, callback_id: str, progress: int, message: str):
        """Update progress for a callback ID."""
        if callback_id:
            self.progress_callbacks[callback_id] = {
                'progress': progress,
                'message': message,
                'completed': progress >= 100,
                'timestamp': timezone.now().isoformat()
            }
    
    def _apply_invoice_filters(self, queryset, filters: Dict[str, Any]):
        """Apply filters to invoice queryset."""
        if filters.get('date_from'):
            queryset = queryset.filter(data_wystawienia__gte=filters['date_from'])
        
        if filters.get('date_to'):
            queryset = queryset.filter(data_wystawienia__lte=filters['date_to'])
        
        if filters.get('status'):
            queryset = queryset.filter(status__in=filters['status'])
        
        if filters.get('typ_dokumentu'):
            queryset = queryset.filter(typ_dokumentu__in=filters['typ_dokumentu'])
        
        return queryset
    
    def _apply_company_filters(self, queryset, filters: Dict[str, Any]):
        """Apply filters to company queryset."""
        if filters.get('czy_firma') is not None:
            queryset = queryset.filter(czy_firma=filters['czy_firma'])
        
        if filters.get('miasto'):
            queryset = queryset.filter(miejscowosc__icontains=filters['miasto'])
        
        return queryset
    
    def _prepare_invoice_data(self, queryset) -> List[Dict[str, Any]]:
        """Prepare invoice data for export."""
        data = []
        
        for faktura in queryset:
            # Calculate totals
            pozycje = faktura.pozycjafaktury_set.all()
            kwota_netto = sum(p.wartosc_netto for p in pozycje)
            kwota_vat = sum(p.wartosc_vat for p in pozycje)
            kwota_brutto = sum(p.wartosc_brutto for p in pozycje)
            
            row = {
                'numer': faktura.numer,
                'typ_dokumentu': faktura.get_typ_dokumentu_display(),
                'data_wystawienia': faktura.data_wystawienia.strftime('%Y-%m-%d'),
                'data_sprzedazy': faktura.data_sprzedazy.strftime('%Y-%m-%d'),
                'termin_platnosci': faktura.termin_platnosci.strftime('%Y-%m-%d'),
                'nabywca__nazwa': faktura.nabywca.nazwa,
                'nabywca__nip': faktura.nabywca.nip or '',
                'nabywca__miejscowosc': faktura.nabywca.miejscowosc,
                'sprzedawca__nazwa': faktura.sprzedawca.nazwa,
                'sprzedawca__nip': faktura.sprzedawca.nip or '',
                'status': faktura.get_status_display(),
                'sposob_platnosci': faktura.get_sposob_platnosci_display(),
                'waluta': faktura.waluta,
                'kwota_netto': float(kwota_netto),
                'kwota_vat': float(kwota_vat),
                'kwota_brutto': float(kwota_brutto),
                'uwagi': faktura.uwagi or ''
            }
            data.append(row)
        
        return data
    
    def _prepare_company_data(self, queryset) -> List[Dict[str, Any]]:
        """Prepare company data for export."""
        data = []
        
        for kontrahent in queryset:
            row = {
                'nazwa': kontrahent.nazwa,
                'nip': kontrahent.nip or '',
                'regon': kontrahent.regon or '',
                'ulica': kontrahent.ulica,
                'numer_domu': kontrahent.numer_domu,
                'numer_mieszkania': kontrahent.numer_mieszkania or '',
                'kod_pocztowy': kontrahent.kod_pocztowy,
                'miejscowosc': kontrahent.miejscowosc,
                'kraj': kontrahent.kraj,
                'czy_firma': 'Firma' if kontrahent.czy_firma else 'Osoba fizyczna',
                'email': kontrahent.email or '',
                'telefon': kontrahent.telefon or '',
                'dodatkowy_opis': kontrahent.dodatkowy_opis or ''
            }
            data.append(row)
        
        return data
    
    def _prepare_product_data(self, queryset) -> List[Dict[str, Any]]:
        """Prepare product data for export."""
        data = []
        
        for produkt in queryset:
            # Calculate brutto price
            vat_rate = float(produkt.vat) / 100 if produkt.vat.isdigit() else 0
            cena_brutto = float(produkt.cena_netto) * (1 + vat_rate)
            
            row = {
                'nazwa': produkt.nazwa,
                'jednostka': produkt.jednostka,
                'cena_netto': float(produkt.cena_netto),
                'vat': f"{produkt.vat}%",
                'cena_brutto': cena_brutto
            }
            data.append(row)
        
        return data
    
    def _export_invoices_excel(self, data: List[Dict], user: User) -> HttpResponse:
        """Export invoices to Excel format."""
        if not Workbook:
            raise ImportError("openpyxl is required for Excel export")
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Faktury"
        
        # Add headers
        headers = list(self.INVOICE_HEADERS_PL.values())
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_data in data:
            row = [row_data.get(key, '') for key in self.INVOICE_HEADERS_PL.keys()]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="faktury_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    def _export_invoices_csv(self, data: List[Dict], user: User) -> HttpResponse:
        """Export invoices to CSV format."""
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=list(self.INVOICE_HEADERS_PL.keys()),
            delimiter=';',  # Polish CSV standard
            quoting=csv.QUOTE_ALL
        )
        
        # Write headers in Polish
        writer.writerow(self.INVOICE_HEADERS_PL)
        
        # Write data
        for row in data:
            writer.writerow(row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="faktury_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for proper Polish characters display in Excel
        response.content = '\ufeff' + response.content.decode('utf-8')
        
        return response
    
    def _export_invoices_pdf(self, data: List[Dict], user: User) -> HttpResponse:
        """Export invoices to PDF format."""
        if not colors:
            raise ImportError("reportlab is required for PDF export")
            
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Content
        story = []
        
        # Title
        title = Paragraph("Lista Faktur", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Table data
        table_data = []
        
        # Headers
        headers = ['Numer', 'Data wyst.', 'Nabywca', 'Status', 'Kwota brutto']
        table_data.append(headers)
        
        # Data rows
        for row in data:
            table_row = [
                row['numer'],
                row['data_wystawienia'],
                row['nabywca__nazwa'][:30] + '...' if len(row['nabywca__nazwa']) > 30 else row['nabywca__nazwa'],
                row['status'],
                f"{row['kwota_brutto']:.2f} {row['waluta']}"
            ]
            table_data.append(table_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="faktury_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
    
    def _export_invoices_json(self, data: List[Dict], user: User) -> HttpResponse:
        """Export invoices to JSON format."""
        export_data = {
            'export_date': timezone.now().isoformat(),
            'user': user.username,
            'data_type': 'invoices',
            'count': len(data),
            'invoices': data
        }
        
        json_data = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        response = HttpResponse(json_data, content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="faktury_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        return response
    
    def _export_companies_excel(self, data: List[Dict], user: User) -> HttpResponse:
        """Export companies to Excel format."""
        if not Workbook:
            raise ImportError("openpyxl is required for Excel export")
        wb = Workbook()
        ws = wb.active
        ws.title = "Kontrahenci"
        
        # Add headers
        headers = list(self.COMPANY_HEADERS_PL.values())
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_data in data:
            row = [row_data.get(key, '') for key in self.COMPANY_HEADERS_PL.keys()]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="kontrahenci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    def _export_companies_csv(self, data: List[Dict], user: User) -> HttpResponse:
        """Export companies to CSV format."""
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=list(self.COMPANY_HEADERS_PL.keys()),
            delimiter=';',
            quoting=csv.QUOTE_ALL
        )
        
        # Write headers in Polish
        writer.writerow(self.COMPANY_HEADERS_PL)
        
        # Write data
        for row in data:
            writer.writerow(row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="kontrahenci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for proper Polish characters display in Excel
        response.content = '\ufeff' + response.content.decode('utf-8')
        
        return response
    
    def _export_companies_pdf(self, data: List[Dict], user: User) -> HttpResponse:
        """Export companies to PDF format."""
        if not colors:
            raise ImportError("reportlab is required for PDF export")
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        
        # Content
        story = []
        
        # Title
        title = Paragraph("Lista Kontrahentów", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Table data
        table_data = []
        
        # Headers
        headers = ['Nazwa', 'NIP', 'Miejscowość', 'Typ', 'Email']
        table_data.append(headers)
        
        # Data rows
        for row in data:
            table_row = [
                row['nazwa'][:25] + '...' if len(row['nazwa']) > 25 else row['nazwa'],
                row['nip'],
                row['miejscowosc'],
                row['czy_firma'],
                row['email'][:25] + '...' if len(row['email']) > 25 else row['email']
            ]
            table_data.append(table_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="kontrahenci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        return response
    
    def _export_companies_json(self, data: List[Dict], user: User) -> HttpResponse:
        """Export companies to JSON format."""
        export_data = {
            'export_date': timezone.now().isoformat(),
            'user': user.username,
            'data_type': 'companies',
            'count': len(data),
            'companies': data
        }
        
        json_data = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        response = HttpResponse(json_data, content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="kontrahenci_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        return response
    
    def _export_products_excel(self, data: List[Dict], user: User) -> HttpResponse:
        """Export products to Excel format."""
        if not Workbook:
            raise ImportError("openpyxl is required for Excel export")
        wb = Workbook()
        ws = wb.active
        ws.title = "Produkty"
        
        # Add headers
        headers = list(self.PRODUCT_HEADERS_PL.values())
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_data in data:
            row = [row_data.get(key, '') for key in self.PRODUCT_HEADERS_PL.keys()]
            ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="produkty_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    def _export_products_csv(self, data: List[Dict], user: User) -> HttpResponse:
        """Export products to CSV format."""
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=list(self.PRODUCT_HEADERS_PL.keys()),
            delimiter=';',
            quoting=csv.QUOTE_ALL
        )
        
        # Write headers in Polish
        writer.writerow(self.PRODUCT_HEADERS_PL)
        
        # Write data
        for row in data:
            writer.writerow(row)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="produkty_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # Add BOM for proper Polish characters display in Excel
        response.content = '\ufeff' + response.content.decode('utf-8')
        
        return response
    
    def _export_products_json(self, data: List[Dict], user: User) -> HttpResponse:
        """Export products to JSON format."""
        export_data = {
            'export_date': timezone.now().isoformat(),
            'user': user.username,
            'data_type': 'products',
            'count': len(data),
            'products': data
        }
        
        json_data = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        response = HttpResponse(json_data, content_type='application/json; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="produkty_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json"'
        
        return response
    
    def _parse_csv_file(self, file_data: BinaryIO) -> List[Dict[str, Any]]:
        """Parse CSV file and return data."""
        content = file_data.read().decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(content), delimiter=';')
        return list(reader)
    
    def _parse_excel_file(self, file_data: BinaryIO) -> List[Dict[str, Any]]:
        """Parse Excel file and return data."""
        if not pd:
            raise ImportError("pandas is required for Excel import")
        df = pd.read_excel(file_data)
        return df.to_dict('records')
    
    def _parse_json_file(self, file_data: BinaryIO) -> List[Dict[str, Any]]:
        """Parse JSON file and return data."""
        content = file_data.read().decode('utf-8')
        data = json.loads(content)
        
        # Handle different JSON structures
        if isinstance(data, dict):
            if 'invoices' in data:
                return data['invoices']
            elif 'companies' in data:
                return data['companies']
            elif 'products' in data:
                return data['products']
            else:
                return [data]
        elif isinstance(data, list):
            return data
        else:
            raise ValueError("Nieprawidłowa struktura pliku JSON")
    
    def _validate_import_data(self, data: List[Dict], data_type: str) -> Dict[str, List[str]]:
        """Validate import data and return errors/warnings."""
        errors = []
        warnings = []
        
        if not data:
            errors.append("Plik nie zawiera danych do importu")
            return {'errors': errors, 'warnings': warnings}
        
        # Validate based on data type
        if data_type == 'invoices':
            errors, warnings = self._validate_invoice_data(data)
        elif data_type == 'companies':
            errors, warnings = self._validate_company_data(data)
        elif data_type == 'products':
            errors, warnings = self._validate_product_data(data)
        else:
            errors.append(f"Nieobsługiwany typ danych: {data_type}")
        
        return {'errors': errors, 'warnings': warnings}
    
    def _validate_invoice_data(self, data: List[Dict]) -> tuple:
        """Validate invoice import data."""
        errors = []
        warnings = []
        
        required_fields = ['numer', 'data_wystawienia', 'nabywca__nazwa']
        
        for i, row in enumerate(data, 1):
            # Check required fields
            for field in required_fields:
                if not row.get(field):
                    errors.append(f"Wiersz {i}: Brak wymaganego pola '{field}'")
            
            # Validate dates
            date_fields = ['data_wystawienia', 'data_sprzedazy', 'termin_platnosci']
            for field in date_fields:
                if row.get(field):
                    try:
                        datetime.strptime(row[field], '%Y-%m-%d')
                    except ValueError:
                        errors.append(f"Wiersz {i}: Nieprawidłowy format daty w polu '{field}'")
            
            # Validate amounts
            amount_fields = ['kwota_netto', 'kwota_vat', 'kwota_brutto']
            for field in amount_fields:
                if row.get(field):
                    try:
                        float(row[field])
                    except (ValueError, TypeError):
                        warnings.append(f"Wiersz {i}: Nieprawidłowa kwota w polu '{field}'")
        
        return errors, warnings
    
    def _validate_company_data(self, data: List[Dict]) -> tuple:
        """Validate company import data."""
        errors = []
        warnings = []
        
        required_fields = ['nazwa', 'miejscowosc']
        
        for i, row in enumerate(data, 1):
            # Check required fields
            for field in required_fields:
                if not row.get(field):
                    errors.append(f"Wiersz {i}: Brak wymaganego pola '{field}'")
            
            # Validate NIP format
            if row.get('nip'):
                nip = row['nip'].replace('-', '').replace(' ', '')
                if not nip.isdigit() or len(nip) != 10:
                    warnings.append(f"Wiersz {i}: Nieprawidłowy format NIP")
            
            # Validate postal code
            if row.get('kod_pocztowy'):
                if not row['kod_pocztowy'].match(r'^\d{2}-\d{3}$'):
                    warnings.append(f"Wiersz {i}: Nieprawidłowy format kodu pocztowego")
        
        return errors, warnings
    
    def _validate_product_data(self, data: List[Dict]) -> tuple:
        """Validate product import data."""
        errors = []
        warnings = []
        
        required_fields = ['nazwa', 'jednostka', 'cena_netto', 'vat']
        
        for i, row in enumerate(data, 1):
            # Check required fields
            for field in required_fields:
                if not row.get(field):
                    errors.append(f"Wiersz {i}: Brak wymaganego pola '{field}'")
            
            # Validate price
            if row.get('cena_netto'):
                try:
                    price = float(row['cena_netto'])
                    if price < 0:
                        warnings.append(f"Wiersz {i}: Cena nie może być ujemna")
                except (ValueError, TypeError):
                    errors.append(f"Wiersz {i}: Nieprawidłowa cena netto")
            
            # Validate VAT rate
            if row.get('vat'):
                vat = str(row['vat']).replace('%', '')
                if vat not in ['0', '5', '8', '23', 'zw']:
                    warnings.append(f"Wiersz {i}: Nieprawidłowa stawka VAT")
        
        return errors, warnings
    
    def _import_validated_data(self, user: User, data: List[Dict], data_type: str, progress_callback_id: str) -> Dict[str, int]:
        """Import validated data to database."""
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        total_rows = len(data)
        
        with transaction.atomic():
            for i, row in enumerate(data):
                try:
                    if data_type == 'invoices':
                        result = self._import_invoice_row(user, row)
                    elif data_type == 'companies':
                        result = self._import_company_row(user, row)
                    elif data_type == 'products':
                        result = self._import_product_row(user, row)
                    else:
                        continue
                    
                    if result == 'imported':
                        imported_count += 1
                    elif result == 'updated':
                        updated_count += 1
                    else:
                        skipped_count += 1
                    
                    # Update progress
                    progress = 60 + (i + 1) / total_rows * 30  # 60-90% range
                    self._update_progress(
                        progress_callback_id,
                        int(progress),
                        f"Przetworzono {i + 1} z {total_rows} rekordów..."
                    )
                    
                except Exception as e:
                    logger.error(f"Error importing row {i}: {str(e)}")
                    skipped_count += 1
        
        return {
            'imported_count': imported_count,
            'updated_count': updated_count,
            'skipped_count': skipped_count
        }
    
    def _import_invoice_row(self, user: User, row: Dict[str, Any]) -> str:
        """Import single invoice row."""
        # This is a simplified implementation
        # In a real scenario, you'd need to handle contractor creation/lookup
        # and proper invoice creation with all relationships
        
        # For now, just return 'skipped' as invoice import is complex
        # and requires proper contractor matching
        return 'skipped'
    
    def _import_company_row(self, user: User, row: Dict[str, Any]) -> str:
        """Import single company row."""
        try:
            # Check if company exists
            existing = Kontrahent.objects.filter(
                user=user,
                nazwa=row['nazwa']
            ).first()
            
            company_data = {
                'user': user,
                'nazwa': row['nazwa'],
                'nip': row.get('nip', ''),
                'regon': row.get('regon', ''),
                'ulica': row.get('ulica', ''),
                'numer_domu': row.get('numer_domu', ''),
                'numer_mieszkania': row.get('numer_mieszkania', ''),
                'kod_pocztowy': row.get('kod_pocztowy', ''),
                'miejscowosc': row.get('miejscowosc', ''),
                'kraj': row.get('kraj', 'Polska'),
                'czy_firma': row.get('czy_firma', '').lower() in ['firma', 'true', '1'],
                'email': row.get('email', ''),
                'telefon': row.get('telefon', ''),
                'dodatkowy_opis': row.get('dodatkowy_opis', '')
            }
            
            if existing:
                # Update existing
                for key, value in company_data.items():
                    if key != 'user':
                        setattr(existing, key, value)
                existing.save()
                return 'updated'
            else:
                # Create new
                Kontrahent.objects.create(**company_data)
                return 'imported'
                
        except Exception as e:
            logger.error(f"Error importing company: {str(e)}")
            return 'skipped'
    
    def _import_product_row(self, user: User, row: Dict[str, Any]) -> str:
        """Import single product row."""
        try:
            # Check if product exists
            existing = Produkt.objects.filter(
                user=user,
                nazwa=row['nazwa']
            ).first()
            
            # Clean VAT rate
            vat = str(row.get('vat', '23')).replace('%', '')
            
            product_data = {
                'user': user,
                'nazwa': row['nazwa'],
                'jednostka': row.get('jednostka', 'szt'),
                'cena_netto': Decimal(str(row.get('cena_netto', '0'))),
                'vat': vat
            }
            
            if existing:
                # Update existing
                for key, value in product_data.items():
                    if key != 'user':
                        setattr(existing, key, value)
                existing.save()
                return 'updated'
            else:
                # Create new
                Produkt.objects.create(**product_data)
                return 'imported'
                
        except Exception as e:
            logger.error(f"Error importing product: {str(e)}")
            return 'skipped'
    
    def _restore_invoices(self, user: User, data: List[Dict]) -> Dict[str, int]:
        """Restore invoices from backup data."""
        # Simplified implementation - would need proper restoration logic
        return {'restored': 0, 'skipped': len(data)}
    
    def _restore_companies(self, user: User, data: List[Dict]) -> Dict[str, int]:
        """Restore companies from backup data."""
        restored = 0
        skipped = 0
        
        for row in data:
            try:
                result = self._import_company_row(user, row)
                if result in ['imported', 'updated']:
                    restored += 1
                else:
                    skipped += 1
            except:
                skipped += 1
        
        return {'restored': restored, 'skipped': skipped}
    
    def _restore_products(self, user: User, data: List[Dict]) -> Dict[str, int]:
        """Restore products from backup data."""
        restored = 0
        skipped = 0
        
        for row in data:
            try:
                result = self._import_product_row(user, row)
                if result in ['imported', 'updated']:
                    restored += 1
                else:
                    skipped += 1
            except:
                skipped += 1
        
        return {'restored': restored, 'skipped': skipped}


# Global service instance
data_export_import_service = DataExportImportService()
"""
API views for data export and import functionality.
Provides REST endpoints for exporting and importing data with progress tracking.
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import json
import uuid
import logging
from datetime import datetime
from typing import Dict, Any

from faktury.services.data_export_import_service import data_export_import_service

logger = logging.getLogger(__name__)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_invoices_api(request):
    """
    Export invoices API endpoint.
    
    POST data:
    {
        "format": "excel|csv|pdf|json",
        "filters": {
            "date_from": "2025-01-01",
            "date_to": "2025-12-31",
            "status": ["wystawiona", "oplacona"],
            "typ_dokumentu": ["FV"]
        }
    }
    """
    try:
        data = json.loads(request.body) if request.body else {}
        
        format_type = data.get('format', 'excel')
        filters = data.get('filters', {})
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Validate format
        if format_type not in ['excel', 'csv', 'pdf', 'json']:
            return Response(
                {'error': 'Nieobsługiwany format eksportu'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Convert date strings to date objects
        if 'date_from' in filters and isinstance(filters['date_from'], str):
            try:
                filters['date_from'] = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
            except ValueError:
                del filters['date_from']
        
        if 'date_to' in filters and isinstance(filters['date_to'], str):
            try:
                filters['date_to'] = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
            except ValueError:
                del filters['date_to']
        
        # Start export
        response = data_export_import_service.export_invoices(
            user=request.user,
            format_type=format_type,
            filters=filters,
            progress_callback_id=progress_id
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in export invoices API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas eksportu faktur'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_companies_api(request):
    """
    Export companies API endpoint.
    """
    try:
        data = json.loads(request.body) if request.body else {}
        
        format_type = data.get('format', 'excel')
        filters = data.get('filters', {})
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Validate format
        if format_type not in ['excel', 'csv', 'pdf', 'json']:
            return Response(
                {'error': 'Nieobsługiwany format eksportu'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Start export
        response = data_export_import_service.export_companies(
            user=request.user,
            format_type=format_type,
            filters=filters,
            progress_callback_id=progress_id
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in export companies API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas eksportu kontrahentów'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_products_api(request):
    """
    Export products API endpoint.
    """
    try:
        data = json.loads(request.body) if request.body else {}
        
        format_type = data.get('format', 'excel')
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Validate format
        if format_type not in ['excel', 'csv', 'json']:
            return Response(
                {'error': 'Nieobsługiwany format eksportu'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Start export
        response = data_export_import_service.export_products(
            user=request.user,
            format_type=format_type,
            progress_callback_id=progress_id
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in export products API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas eksportu produktów'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_backup_api(request):
    """
    Create complete data backup API endpoint.
    """
    try:
        data = json.loads(request.body) if request.body else {}
        
        include_files = data.get('include_files', False)
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Start backup creation
        response = data_export_import_service.create_backup(
            user=request.user,
            include_files=include_files,
            progress_callback_id=progress_id
        )
        
        return response
        
    except Exception as e:
        logger.error(f"Error in create backup API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas tworzenia kopii zapasowej'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_data_api(request):
    """
    Import data from file API endpoint.
    
    Form data:
    - file: File to import
    - data_type: 'invoices', 'companies', 'products'
    - format_type: 'csv', 'excel', 'json'
    """
    try:
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Nie przesłano pliku'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        uploaded_file = request.FILES['file']
        data_type = request.POST.get('data_type', 'companies')
        format_type = request.POST.get('format_type', 'csv')
        
        # Validate parameters
        if data_type not in ['invoices', 'companies', 'products']:
            return Response(
                {'error': 'Nieobsługiwany typ danych'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if format_type not in ['csv', 'excel', 'json']:
            return Response(
                {'error': 'Nieobsługiwany format pliku'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Start import
        result = data_export_import_service.import_data(
            user=request.user,
            file_data=uploaded_file,
            data_type=data_type,
            format_type=format_type,
            progress_callback_id=progress_id
        )
        
        if result['success']:
            return Response(result, status=status.HTTP_200_OK)
        else:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        logger.error(f"Error in import data API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas importu danych'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def restore_backup_api(request):
    """
    Restore data from backup file API endpoint.
    """
    try:
        if 'file' not in request.FILES:
            return Response(
                {'error': 'Nie przesłano pliku kopii zapasowej'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        backup_file = request.FILES['file']
        
        # Validate file type
        if not backup_file.name.endswith('.zip'):
            return Response(
                {'error': 'Plik kopii zapasowej musi być w formacie ZIP'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate progress callback ID
        progress_id = str(uuid.uuid4())
        
        # Start restore
        result = data_export_import_service.restore_backup(
            user=request.user,
            backup_file=backup_file,
            progress_callback_id=progress_id
        )
        
        if result['success']:
            return Response(result, status=status.HTTP_200_OK)
        else:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        logger.error(f"Error in restore backup API: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas przywracania kopii zapasowej'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_progress_api(request, progress_id):
    """
    Get export/import progress status.
    """
    try:
        progress = data_export_import_service.get_export_progress(progress_id)
        return Response(progress, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error getting export progress: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas pobierania statusu'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_templates_api(request):
    """
    Get export templates for different data types.
    """
    try:
        data_type = request.GET.get('type', 'companies')
        format_type = request.GET.get('format', 'csv')
        
        if data_type not in ['invoices', 'companies', 'products']:
            return Response(
                {'error': 'Nieobsługiwany typ danych'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if format_type not in ['csv', 'excel']:
            return Response(
                {'error': 'Nieobsługiwany format szablonu'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generate empty template
        if data_type == 'companies':
            template_data = [data_export_import_service.COMPANY_HEADERS_PL]
        elif data_type == 'products':
            template_data = [data_export_import_service.PRODUCT_HEADERS_PL]
        else:
            template_data = [data_export_import_service.INVOICE_HEADERS_PL]
        
        # Create template file
        if format_type == 'csv':
            response = _create_csv_template(template_data[0], data_type)
        else:
            response = _create_excel_template(template_data[0], data_type)
        
        return response
        
    except Exception as e:
        logger.error(f"Error creating export template: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas tworzenia szablonu'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_statistics_api(request):
    """
    Get export statistics for the user.
    """
    try:
        from faktury.models import Faktura, Kontrahent, Produkt
        from django.db.models import Count
        
        stats = {
            'total_invoices': Faktura.objects.filter(user=request.user).count(),
            'total_companies': Kontrahent.objects.filter(user=request.user).count(),
            'total_products': Produkt.objects.filter(user=request.user).count(),
            'invoice_types': list(
                Faktura.objects.filter(user=request.user)
                .values('typ_dokumentu')
                .annotate(count=Count('id'))
            ),
            'invoice_statuses': list(
                Faktura.objects.filter(user=request.user)
                .values('status')
                .annotate(count=Count('id'))
            ),
            'company_types': list(
                Kontrahent.objects.filter(user=request.user)
                .values('czy_firma')
                .annotate(count=Count('id'))
            )
        }
        
        return Response(stats, status=status.HTTP_200_OK)
        
    except Exception as e:
        logger.error(f"Error getting export statistics: {str(e)}")
        return Response(
            {'error': 'Wystąpił błąd podczas pobierania statystyk'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@method_decorator(csrf_exempt, name='dispatch')
@method_decorator(login_required, name='dispatch')
class BulkOperationsView(View):
    """
    Handle bulk operations on data.
    """
    
    def post(self, request):
        """Handle bulk operations requests."""
        try:
            data = json.loads(request.body)
            operation = data.get('operation')
            item_ids = data.get('item_ids', [])
            data_type = data.get('data_type', 'companies')
            
            if not operation or not item_ids:
                return JsonResponse({
                    'error': 'Brak operacji lub elementów do przetworzenia'
                }, status=400)
            
            if operation == 'delete':
                result = self._bulk_delete(request.user, item_ids, data_type)
            elif operation == 'export':
                result = self._bulk_export(request.user, item_ids, data_type)
            elif operation == 'update_status':
                new_status = data.get('new_status')
                result = self._bulk_update_status(request.user, item_ids, data_type, new_status)
            else:
                return JsonResponse({
                    'error': f'Nieobsługiwana operacja: {operation}'
                }, status=400)
            
            return JsonResponse(result)
            
        except Exception as e:
            logger.error(f"Error in bulk operations: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    def _bulk_delete(self, user, item_ids, data_type):
        """Perform bulk delete operation."""
        try:
            if data_type == 'invoices':
                from faktury.models import Faktura
                deleted_count = Faktura.objects.filter(
                    user=user, id__in=item_ids
                ).delete()[0]
            elif data_type == 'companies':
                from faktury.models import Kontrahent
                deleted_count = Kontrahent.objects.filter(
                    user=user, id__in=item_ids
                ).delete()[0]
            elif data_type == 'products':
                from faktury.models import Produkt
                deleted_count = Produkt.objects.filter(
                    user=user, id__in=item_ids
                ).delete()[0]
            else:
                return {'error': 'Nieobsługiwany typ danych'}
            
            return {
                'success': True,
                'deleted_count': deleted_count,
                'message': f'Usunięto {deleted_count} elementów'
            }
            
        except Exception as e:
            return {'error': f'Błąd podczas usuwania: {str(e)}'}
    
    def _bulk_export(self, user, item_ids, data_type):
        """Perform bulk export operation."""
        try:
            # Create filtered export based on selected items
            filters = {'ids': item_ids}
            
            if data_type == 'invoices':
                response = data_export_import_service.export_invoices(
                    user=user,
                    format_type='excel',
                    filters=filters
                )
            elif data_type == 'companies':
                response = data_export_import_service.export_companies(
                    user=user,
                    format_type='excel',
                    filters=filters
                )
            elif data_type == 'products':
                response = data_export_import_service.export_products(
                    user=user,
                    format_type='excel'
                )
            else:
                return {'error': 'Nieobsługiwany typ danych'}
            
            return {
                'success': True,
                'message': 'Eksport rozpoczęty',
                'download_url': '/download/bulk_export.xlsx'  # Placeholder
            }
            
        except Exception as e:
            return {'error': f'Błąd podczas eksportu: {str(e)}'}
    
    def _bulk_update_status(self, user, item_ids, data_type, new_status):
        """Perform bulk status update operation."""
        try:
            if data_type == 'invoices' and new_status:
                from faktury.models import Faktura
                updated_count = Faktura.objects.filter(
                    user=user, id__in=item_ids
                ).update(status=new_status)
                
                return {
                    'success': True,
                    'updated_count': updated_count,
                    'message': f'Zaktualizowano status {updated_count} faktur'
                }
            else:
                return {'error': 'Operacja dostępna tylko dla faktur'}
                
        except Exception as e:
            return {'error': f'Błąd podczas aktualizacji: {str(e)}'}


def _create_csv_template(headers, data_type):
    """Create CSV template file."""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=list(headers.keys()),
        delimiter=';',
        quoting=csv.QUOTE_ALL
    )
    
    # Write headers in Polish
    writer.writerow(headers)
    
    # Add one example row
    if data_type == 'companies':
        example_row = {
            'nazwa': 'Przykładowa Firma Sp. z o.o.',
            'nip': '1234567890',
            'regon': '123456789',
            'ulica': 'Przykładowa',
            'numer_domu': '1',
            'numer_mieszkania': '',
            'kod_pocztowy': '00-001',
            'miejscowosc': 'Warszawa',
            'kraj': 'Polska',
            'czy_firma': 'Firma',
            'email': 'kontakt@przyklad.pl',
            'telefon': '123456789',
            'dodatkowy_opis': 'Przykładowy opis'
        }
    elif data_type == 'products':
        example_row = {
            'nazwa': 'Przykładowy produkt',
            'jednostka': 'szt',
            'cena_netto': '100.00',
            'vat': '23%',
            'cena_brutto': '123.00'
        }
    else:
        example_row = {}
    
    if example_row:
        writer.writerow(example_row)
    
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="szablon_{data_type}.csv"'
    
    # Add BOM for proper Polish characters display in Excel
    response.content = '\ufeff' + response.content.decode('utf-8')
    
    return response


def _create_excel_template(headers, data_type):
    """Create Excel template file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel templates")
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Szablon {data_type}"
    
    # Add headers
    header_values = list(headers.values())
    ws.append(header_values)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(header_values, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add example row
    if data_type == 'companies':
        example_row = [
            'Przykładowa Firma Sp. z o.o.',
            '1234567890',
            '123456789',
            'Przykładowa',
            '1',
            '',
            '00-001',
            'Warszawa',
            'Polska',
            'Firma',
            'kontakt@przyklad.pl',
            '123456789',
            'Przykładowy opis'
        ]
    elif data_type == 'products':
        example_row = [
            'Przykładowy produkt',
            'szt',
            100.00,
            '23%',
            123.00
        ]
    else:
        example_row = []
    
    if example_row:
        ws.append(example_row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="szablon_{data_type}.xlsx"'
    
    return response